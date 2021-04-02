from common import *
from reg_tree import *
import copy

def extract_info(top,rb,gen_inst=False):

    block_col = ["reg_block", "address_width", "data_width", "base_address", "range","parameters", \
                "abbre_name", "bus_reg_out", "wdata_bp", "addr_unit_bits","path"]
    reg_col   = ["register", "offset", "size", "access", "reset_value", \
                "description", "abbre_name", "index", "", "path","statement"]
    field_col = ["field", "width", "bits", "access", "reset_value", \
                "description", "abbre_name", "rtl_io_access", "", "",""]

    logger.info("Start to collect excel information!")
    collect_blocks = block("AIPU")
    for sht_name in rb.sheetnames:
        if(sht_name != "Home" or sht_name != "Help"):
            sht = rb[sht_name]
            if(sht['A1'].value == "Register Specification"):
                for i in range(1,sht.max_row):
                    row = list(sht.rows)[i]
                    if(row[0].value == "reg_block"):
                        i = i+1
                        row = list(sht.rows)[i]
                        if row[block_col.index("reg_block")].value == None:
                            logger.warning("Detect a unnamed block, the follwing registers of this sheet will append to last block")
                            continue
                        #---- BLOCK: Format variable name
                        abbre_name = row[block_col.index("abbre_name")].value
                        if abbre_name == None:
                            abbre_name = row[block_col.index("reg_block")].value
                        abbre_name=abbre_name.strip().replace(' ','_')
                        tmp_name = abbre_name
                        suffix=0
                        while tmp_name in collect_blocks.blocks:
                            logger.error("Block:{} has been created in the top block before. please modify a uniqe abbre_name".format(abbre_name))
                            suffix+=1
                            continue
                            tmp_name=abbre_name+"_{}".format(suffix)
                        abbre_name = tmp_name
                        if len(abbre_name) > 15:
                            logger.error("abbre_name".format(abbre_name))

                        #----BLOCK: Extract block info
                        new_block = block("new_block")
                        new_block.name = abbre_name
                        new_block.full_name = row[block_col.index("reg_block")].value.strip()
                        new_block.address_width = row[block_col.index("address_width")].value
                        new_block.data_width = row[block_col.index("data_width")].value
                        new_block.base_address = row[block_col.index("base_address")].value or '0x0'
                        new_block.range = row[block_col.index("range")].value or 2**int(new_block.address_width)
                        new_block.parameters = row[block_col.index("parameters")].value or ''
                        new_block.bus_reg_out = row[block_col.index("bus_reg_out")].value or False
                        new_block.wdata_bp = row[block_col.index("wdata_bp")].value or '0'
                        new_block.addr_unit_bits = row[block_col.index("addr_unit_bits")].value or '1'
                        new_block.path = row[block_col.index("path")].value or ''
                        collect_blocks.add_block(new_block)
                        #---- BLOCK:  Check info & format info
                        if new_block.address_width == None:
                            logger.error("Required information(address_width) is error in reg block: {} ".format(new_block.full_name))
                        if new_block.data_width == None:
                            logger.error("Required information(data_width) is error in reg block: {} ".format(new_block.full_name))
                        new_block.name = new_block.name.lower()
                        new_block.address_width = int(new_block.address_width)
                        new_block.data_width = int(new_block.data_width)
                        new_block.bus_reg_out = bool(new_block.bus_reg_out.lower()=='true')
                        new_block.wdata_bp = int(new_block.wdata_bp)
                        new_block.addr_unit_bits = int(new_block.addr_unit_bits)
                        new_block.range = hex(eval(str(new_block.range)))

                        logger.info("Found a reg block: {},{}".format(new_block.full_name,new_block.name))

                    if(row[0].value == "register"):
                        i = i+1
                        row = list(sht.rows)[i]
                        if row[reg_col.index("register")].value == None: 
                            logger.warning("Detect a named register,Skip the extraction of this register")
                            continue
                        #---- Format variable name
                        abbre_name = row[reg_col.index("abbre_name")].value
                        if abbre_name == None: 
                            abbre_name = row[reg_col.index("register")].value
                        if abbre_name == None: 
                            abbre_name = "unnamed_register"
                        abbre_name=abbre_name.strip().replace(' ','_')
                        suffix=0
                        tmp_name = abbre_name
                        while tmp_name in new_block.regs:
                            suffix+=1
                            tmp_name=abbre_name+"_{}".format(suffix)
                        abbre_name = tmp_name
                        #---- Extrace register info
                        new_reg = register("new_reg")
                        new_reg.name = abbre_name
                        new_reg.full_name = row[reg_col.index("register")].value.strip()
                        new_reg.offset = row[reg_col.index("offset")].value
                        new_reg.size = row[reg_col.index("size")].value or new_block.data_width
                        new_reg.access = row[reg_col.index("access")].value or 'RW'
                        new_reg.reset_value = row[reg_col.index("reset_value")].value or '0x0'
                        new_reg.description = row[reg_col.index("description")].value or ''
                        new_reg.index = row[reg_col.index("index")].value or '0'
                        new_reg.path = row[reg_col.index("path")].value or ''
                        new_reg.statement = row[reg_col.index("statement")].value or ''
                        new_block.add_reg(new_reg)
                        #---- REG:  Check info & format info
                        if new_reg.offset  == None:
                            logger.error("Required information(offset) is error in reg: {} ".format(new_reg.full_name))
                        new_reg.name = new_reg.name.lower()
                        new_reg.size = int(new_reg.size)
                        new_reg.index = int(new_reg.index)

                        if i == sht.max_row-2:
                            break
                        else:
                            i = i+2
                            row = list(sht.rows)[i]
                        while(row[0].value != None and row[0].value != "register" and row[0].value != "reg_lock"):
                            #---- Format variable name
                            abbre_name = row[field_col.index("abbre_name")].value
                            if abbre_name == None:
                                abbre_name = row[field_col.index("field")].value
                            abbre_name=abbre_name.strip().replace(' ','_')
                            suffix=0
                            tmp_name = abbre_name
                            if abbre_name == 'Reserved':
                                abbre_name = 'RSD'
                                tmp_name = 'RSD_0'
                            while tmp_name in new_reg.fields:
                                suffix+=1
                                tmp_name=abbre_name+"_{}".format(suffix)
                            abbre_name = tmp_name
                            #---- Extrace field info
                            new_field = field("new_field")
                            new_field.name = abbre_name
                            new_field.full_name = row[field_col.index("field")].value.strip()
                            new_field.width = row[field_col.index("width")].value
                            new_field.bits = row[field_col.index("bits")].value
                            new_field.access = row[field_col.index("access")].value
                            new_field.reset_value = row[field_col.index("reset_value")].value
                            new_field.description = row[field_col.index("description")].value or ''
                            new_field.rtl_io_access = row[field_col.index("rtl_io_access")].value or "HNONE"
                            new_reg.add_field(new_field)
                            logger.debug("In register:{}, Add a new field:{} ".format(new_reg.name,new_field.name))
                            
                            #---- REG:  Check info & format info
                            if new_field.width  == None:
                                logger.error("Required information(width) is error in reg field: {}.{}".format(new_reg.full_name,new_field.full_name))
                            if new_field.bits  == None:
                                logger.error("Required information(bits) is error in reg field: {}.{} ".format(new_reg.full_name,new_field.full_name))
                            if new_field.access  == None:
                                logger.error("Required information(access) is error in reg field: {}.{} ".format(new_reg.full_name,new_field.full_name))
                            if new_field.reset_value  == None:
                                logger.error("Required information(reset_value) is error in reg field: {}.{} ".format(new_reg.full_name,new_field.full_name))
                            new_field.name = new_field.name.lower()
                            new_field.width = int(new_field.width)
                            new_field.bits = str(new_field.bits).replace('：',':') #replace ":" full to half width
                            new_field.update_bits()

                            if '[' in new_field.reset_value or ']' in new_field.reset_value:
                                new_field.reset_value = new_field.reset_value.replace('[','').replace(']','')
                            if new_field.reset_value.strip() =='-':
                                new_field.reset_value = '0x0'
                                new_field.reset_mask = '0x0'
                                new_field.no_reset = '1'
                            else:
                                new_field.reset_mask = hex(2**new_field.width-1)
                            if 'W' in new_field.rtl_io_access:
                                new_field.volatile = 'true'
                            else:
                                new_field.volatile = 'false'

                            if i == sht.max_row-1:
                                break
                            else:
                                i = i+1
                                row = list(sht.rows)[i]

    # instantiate block in top block
    sht = rb["Home"]
    top.pjname = sht["J2"].value
    abbre_name = sht["J8"].value
    if abbre_name == None:
        abbre_name = sht["J3"].value
    abbre_name=abbre_name.strip().replace(' ','_')
    top.name = abbre_name
    top.full_name = sht["J3"].value
    top.address_width = sht["J4"].value
    top.data_width = sht["J5"].value
    top.base_address = sht["J6"].value
    top.addr_unit_bits = sht["J7"].value
    top.range = sht["J9"].value or 2**int(top.address_width)
    top.range = hex(eval(str(top.range)))

    new_sub = sht["I16"]
    while new_sub.value:
        inst = dict()
        inst["block_name"] = new_sub.value
        inst["block_instance"] = new_sub.offset(0,1).value
        inst["parameters"] = new_sub.offset(0,2).value or ''
        inst["base_address"] = new_sub.offset(0,3).value or ''
        inst["path"] = new_sub.offset(0,4).value or ''
        top.sub_block_info.append(inst)
        new_sub = new_sub.offset(1,0)
    
    for block_name in collect_blocks.blocks:
        top.add_block(collect_blocks.blocks[block_name])

    logger.info("Extract excel information to Reg tree done!")


    if gen_inst == True:
        inst_top=block2inst(top)
        return top,inst_top
    else:
        return top,None


def block2inst(top):

    inst_top = copy.deepcopy(top)
    inst_top.blocks.clear()
    
    import math
    addr_len_hex = math.ceil(int(inst_top.address_width)/4)

    for inst in top.sub_block_info:
        inst_block = copy.deepcopy(top.blocks[inst["block_name"]])
        inst_block.name = inst["block_instance"].strip().replace(' ','_')
        if inst['parameters'] != '':
            inst_block.parameters = inst["parameters"].replace(' ','').replace('，',',')
        if inst['base_address'] != '':
            inst_block.base_address = inst["base_address"]
        if inst['path'] != '':
            inst_block.path  = inst["path"]

        if inst['parameters'] != '':
            parameter = dict()
            for single in inst_block.parameters.split(','):
                para=single.split('=')
                parameter[para[0]]=para[1]

            #  block parameter replace
            for val in ['full_name','path']:
                if type(val) == type('') : 
                    for para,para_inst in parameter.items():
                        ex_str = "{obj}.{val} = {obj}.{val}.replace('[{para_name}]','{para_inst}')"
                        ex_str = ex_str.format(obj='inst_block',val=val,para_name=para,para_inst=para_inst)
                        exec(ex_str)

            #  reg parameter replace
            for reg_name in inst_block.regs:
                reg = inst_block.regs[reg_name]
                for val in ['full_name','name','description','statement','path']:
                    if type(val) == type('') : 
                        for para,para_inst in parameter.items():
                            ex_str = "{obj}.{val} = {obj}.{val}.replace('[{para_name}]','{para_inst}')"
                            ex_str = ex_str.format(obj='reg',val=val,para_name=para,para_inst=para_inst)
                            exec(ex_str)

                # field parameter replace
                for field_name in reg.fields:
                    field = reg.fields[field_name]
                    for val in ['full_name','name','reset_value']:
                        if type(val) == type('') : 
                            for para,para_inst in parameter.items():
                                ex_str = "{obj}.{val} = {obj}.{val}.replace('[{para_name}]','{para_inst}')"
                                ex_str = ex_str.format(obj='field',val=val,para_name=para,para_inst=para_inst)
                                exec(ex_str)
                    # reset_value has no '[]'
                    ex_str = "{obj}.{val} = {obj}.{val}.replace('{para_name}','{para_inst}')"
                    ex_str = ex_str.format(obj='field',val='reset_value',para_name=para,para_inst=para_inst)
                    exec(ex_str)

        # fill reg other info from all fields
        for reg_name in inst_block.regs:
            reg = inst_block.regs[reg_name]
            reg.offset = eval(inst_block.base_address) + eval(reg.offset)
            cmd_str = "'{{:0{size}x}}'.format({value})"
            reg.offset = '0x'+eval(cmd_str.format(size=addr_len_hex,value=reg.offset))

            for field_name in reg.fields:
                field = reg.fields[field_name]
                field.reset_value= hex(eval(field.reset_value))

            has_w = 0
            has_w = 0
            has_vlt = 0
            rst_val_bin = ''
            rst_msk_bin = ''
            for field_name in reg.fields:
                f = reg.fields[field_name]
                cmd_str = "'{{:0{size}b}}'.format(int('{value}', 16))"
                rst_val_bin += eval(cmd_str.format(size=f.width,value=f.reset_value))
                rst_msk_bin += eval(cmd_str.format(size=f.width,value=f.reset_mask))
                if 'W' in f.access: 
                    has_w = 1
                if 'R' in f.access: 
                    has_r = 1
                if  f.volatile == 'true': 
                    has_vlt = 1
            # TODO: check bit order and check all bits width
            reg.reset_value = '0x'+'{:08x}'.format(int(rst_val_bin, 2))
            reg.reset_mask = '0x'+'{:08x}'.format(int(rst_msk_bin, 2))
            if has_w == 0 and has_r == 1:
                reg.access = "RO"
            elif has_w == 1 and has_r == 0:
                reg.access = "WO"
            else: 
                reg.access = "RW"
            if has_vlt == 1:
                reg.volatile = "true"
            else:
                reg.volatile = "false"

        inst_top.add_block(inst_block)

    logger.debug("Extract excel information to Reg tree done(instance)!")
    return inst_top


def get_demo_reg_tree():

    import argparse
    from openpyxl import load_workbook

    description = "This script is designed to parse the Register file, And extract information to a Reg_tree class" 
    parser = argparse.ArgumentParser(description=description)  
    # parser.add_argument('-f','--excel_file',type=str, default='c:/Users/kyltao01/Desktop/Depositary/Teamp Project/Reg_parse/Reg_template.xlsm', help='Indicate a register document file.')
    parser.add_argument('-f','--excel_file',type=str, default='c:/Users/kyltao01/Desktop/Depositary/Teamp Project/Reg_parse/tpc_reg.xlsm', help='Indicate a register document file.')
    args=parser.parse_args()
    filename=args.excel_file
    rb = load_workbook(filename,read_only=False)

    top = block("TOP")
    top,inst_top = extract_info(top,rb,True)
    return top,inst_top 

if __name__ == "__main__":  
    get_demo_reg_tree()
